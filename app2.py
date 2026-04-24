import streamlit as st
import os
import pandas as pd
from openpyxl import Workbook, load_workbook

st.title("Student Entry System")

# =========================
# ✏️ NAME INPUT (AT THE TOP)
# =========================
new_student = st.text_input("Enter Student Name")

# =========================
# 📌 SCHOOL LIST
# =========================
schools = [
    "KNUST SHS",
    "Prempeh College",
    "Opoku Ware School",
    "Yaa Asantewaa Girls"
]

# =========================
# 📌 SHEETS
# =========================
sheets = [
    "KG1 SBA", "KG2 SBA",
    "P1 SBA", "P2 SBA", "P3 SBA", "P4 SBA",
    "P5 SBA", "P6 SBA",
    "JHS 1 SBA", "JHS 2 SBA", "JHS 3 SBA"
]

# =========================
# 🔽 SELECT SCHOOL & SHEET
# =========================
school = st.selectbox("Select School", schools)
sheet_name = st.selectbox("Select Class Sheet", sheets)

# =========================
# 📂 FILE NAME
# =========================
school_file = school.replace(" ", "_").lower() + ".xlsx"

# =========================
# 📁 CREATE FILE IF NOT EXISTS
# =========================
if not os.path.exists(school_file):
    wb = Workbook()
    wb.remove(wb.active)

    for s in sheets:
        ws = wb.create_sheet(title=s)
        ws.append(["Name"])

    wb.save(school_file)

# =========================
# 📄 LOAD EXISTING NAMES
# =========================
try:
    df = pd.read_excel(school_file, sheet_name=sheet_name)
except:
    df = pd.DataFrame(columns=["Name"])

if "Name" not in df.columns:
    df["Name"] = []

name_list = df["Name"].dropna().tolist()

# =========================
# 🔽 SHOW EXISTING NAMES
# =========================
st.selectbox(
    "Existing Students",
    name_list if name_list else ["No names yet"]
)

# =========================
# ➕ ADD BUTTON
# =========================
if st.button("Add Student"):
    if new_student.strip() == "":
        st.warning("Enter a valid name")
    else:
        wb = load_workbook(school_file)
        ws = wb[sheet_name]

        ws.append([new_student])
        wb.save(school_file)

        st.success(f"✅ {new_student} added to {sheet_name}")
        # =========================
# 📊 DISPLAY EDITABLE TABLE (NAME LOCKED)
# =========================

st.subheader(f"{school} - {sheet_name}")

try:
    df = pd.read_excel(school_file, sheet_name=sheet_name)
except:
    df = pd.DataFrame(columns=["Name"])

# Keep original names
original_names = df["Name"].copy() if "Name" in df.columns else []

# Editable table
edited_df = st.data_editor(df, num_rows="dynamic")

# Save button
if st.button("Save Changes"):
    # 🔒 Restore Name column (prevents editing)
    if "Name" in edited_df.columns:
        edited_df["Name"] = original_names

    edited_df.to_excel(
        school_file,
        sheet_name=sheet_name,
        index=False
    )

    st.success("✅ Changes saved (Name column protected)")